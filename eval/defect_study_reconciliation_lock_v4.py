#!/usr/bin/env python3
"""Commit to the reconciliation order before any label exists, then derive it afterwards.

Why this file exists
--------------------
The reviewer's P0-4 asks for the calibration study to be re-run with at least three fully blind
annotators and with the reconciliation "locked in advance, not sorted by the geometric fields".
The second half is the part that killed the previous session's reconciliation: its working note
shows the disputed rows had been pre-sorted by patch geometry computed outside the blinded packets
(whether the finding's file was patched, whether its function was, and the distance to the nearest
changed line). Those are exactly the fields the study exists to test, so a label produced while
reading rows in that order cannot be evidence about that order's subject. The session was held, its
result was read, and it was then excluded. Excluding a registered step after seeing its result is a
worse position than never having run it, which is why the reviewer wants the whole thing redone
rather than patched.

The obvious fix, "write the row order down before labelling", cannot be implemented literally. The
reconciliation set is the set of findings the annotators disagreed on, and that set is unknown until
the sheets come back. What can be fixed in advance is the *rule* that produces the order from the
disputed set, together with the seed it uses. So this file implements a commitment:

    lock    (before labelling)  record the rule id, the seed, the sha256 of this script, and the
                                sha256 of every annotator workbook as issued, having first checked
                                that every label cell in them is empty.
    derive  (after labelling)   read the returned sheets, compute the disputed set, and emit the
                                reconciliation workbook in the order the committed rule produces.

`eval/check_study_blinding_v4.py` then re-derives the order from the committed seed and refuses it
if it does not match, and separately refuses an order whose position is associated with any
geometric field. Neither check trusts this script's output: both recompute.

The rule is a seeded permutation of the disputed finding uids. A permutation is the right object
because it has no relationship to any property of a finding, which is the property being asserted,
and because it is reproducible from the seed alone, so nobody has to be trusted about it later.

    python3 -m eval.defect_study_reconciliation_lock_v4 lock --reviewers A,B,C
    python3 -m eval.defect_study_reconciliation_lock_v4 derive --reviewers A,B,C
"""
from __future__ import annotations
import argparse, hashlib, json, os, sys, time

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
from eval import adjudication_v3_common as C

SYS_ROOT = C.SYS_ROOT
OUT_DIR = os.path.join(SYS_ROOT, "revision-cns-v2", "out")
STUDY_DIR = os.path.join(C.ADJ_DIR, "study-v4")
LOCK = os.path.join(STUDY_DIR, "RECONCILIATION_ORDER_LOCK.json")

# The rule id is part of the commitment. Changing how order_of() works without changing this string
# would let a later order pass a check written against an earlier promise.
RULE_ID = "seeded-permutation-of-finding-uid-v1"
SEED = 20260820

# The axis a reconciliation session is held over. Two annotators can differ on any of the five, but
# the study's number is the root-cause axis, so that is the axis a dispute is defined on.
DISPUTE_AXIS = "root_cause_relation"


def _sha256_file(p: str) -> str:
    h = hashlib.sha256()
    with open(p, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def order_of(uids, seed: int = SEED):
    """The committed rule: a seeded permutation of the disputed finding uids.

    Implemented on hashlib rather than random.shuffle so that it does not depend on the Python
    version's PRNG, which has changed shape before and would silently re-order a study held on a
    different interpreter. Each uid is scored by sha256(seed|uid) and the scores are sorted. That is
    a permutation determined entirely by the seed and the uid set, computable by anyone with both,
    and independent of any property of the finding by construction.
    """
    def score(u: str) -> str:
        return hashlib.sha256(f"{seed}|{u}".encode()).hexdigest()
    return sorted(uids, key=score)


def _label_cells_blank(path: str) -> tuple[bool, int, int]:
    """(all blank, n rows, n filled cells) over the five label axes of a tier-2 workbook.

    A lock taken over a workbook that already carries labels proves nothing, so this is checked
    rather than assumed, and the counts are recorded either way.
    """
    try:
        import openpyxl
    except Exception:
        return (False, 0, -1)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    filled = rows = 0
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        cols = {h: i for i, h in enumerate(header)}
        idx = [cols[a] for a in C.TIER2_LABEL_AXES if a in cols]
        if not idx:
            continue
        for row in it:
            if all(row[i] in (None, "") for i in range(len(header)) if i < len(row)):
                continue
            rows += 1
            for i in idx:
                if i < len(row) and row[i] not in (None, ""):
                    filled += 1
    return (filled == 0, rows, filled)


def cmd_lock(a) -> int:
    os.makedirs(STUDY_DIR, exist_ok=True)
    issued = {}
    for tag in a.reviewers:
        p = os.path.join(a.packages, f"reviewer_{tag}", f"reviewer_{tag}.xlsx")
        if not os.path.isfile(p):
            p = os.path.join(a.packages, f"reviewer_{tag}.xlsx")
        if not os.path.isfile(p):
            print(f"  reviewer {tag}: no workbook at {p}, recorded as not issued")
            issued[tag] = {"path": None, "sha256": None, "blank": None}
            continue
        blank, rows, filled = _label_cells_blank(p)
        issued[tag] = {"path": os.path.relpath(p, SYS_ROOT), "sha256": _sha256_file(p),
                       "rows": rows, "filled_label_cells": filled, "blank": blank}
        print(f"  reviewer {tag}: {rows} rows, {filled} filled label cells, blank={blank}")

    not_blank = [t for t, v in issued.items() if v.get("blank") is False]
    if not_blank and not a.allow_labelled:
        sys.exit("REFUSING TO LOCK: workbooks already carry labels for " + ", ".join(not_blank) +
                 ". A commitment taken after the labels exist is not a commitment. Pass "
                 "--allow-labelled only to record a lock that is explicitly not blind, and expect "
                 "check_study_blinding_v4 to report it as such.")

    payload = {
        "schema_version": "reconciliation-order-lock-v4",
        "purpose": ("commit to the rule that orders the reconciliation session, before the "
                    "disputed set is known, so the order cannot be chosen after the disagreements "
                    "are visible"),
        "rule_id": RULE_ID,
        "seed": a.seed,
        "dispute_axis": DISPUTE_AXIS,
        "script": "eval/defect_study_reconciliation_lock_v4.py",
        "script_sha256": _sha256_file(os.path.abspath(__file__)),
        "reviewers": list(a.reviewers),
        "workbooks_as_issued": issued,
        "all_workbooks_blank_at_lock": all(v.get("blank") for v in issued.values()) if issued else None,
        "locked_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "what_this_does_not_promise": (
            "the disputed set itself is not committed here and cannot be, because it does not "
            "exist until the sheets return. What is committed is the seed and the rule, so the "
            "order derived later is a function of the disputed set alone and of nothing anyone "
            "learned by reading it."),
    }
    payload["content_hash"] = C.content_hash({k: v for k, v in payload.items()
                                              if k != "content_hash"})
    json.dump(payload, open(LOCK, "w", encoding="utf-8"), indent=1, sort_keys=True)
    print("locked -> " + os.path.relpath(LOCK, SYS_ROOT))
    print("  rule  " + RULE_ID + f"  seed {a.seed}")
    print("  hash  " + payload["content_hash"][:16])
    return 0


def _read_sheet(path: str) -> dict:
    """finding_uid -> {axis: value} from a returned tier-2 workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict = {}
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        cols = {h: i for i, h in enumerate(header)}
        uid_col = cols.get("finding_uid", cols.get("packet_id"))
        if uid_col is None or not any(a in cols for a in C.TIER2_LABEL_AXES):
            continue
        for row in it:
            if uid_col >= len(row) or row[uid_col] in (None, ""):
                continue
            out[str(row[uid_col]).strip()] = {
                a: (str(row[cols[a]]).strip() if a in cols and cols[a] < len(row)
                    and row[cols[a]] is not None else "")
                for a in C.TIER2_LABEL_AXES}
    return out


def cmd_derive(a) -> int:
    if not os.path.isfile(LOCK):
        sys.exit(f"no lock at {LOCK}. Run `lock` before labelling, not after.")
    lock = json.load(open(LOCK, encoding="utf-8"))
    if lock["rule_id"] != RULE_ID:
        sys.exit(f"the lock commits to rule {lock['rule_id']} and this script implements "
                 f"{RULE_ID}. Check out the script the lock names rather than re-pointing the lock.")

    sheets = {}
    for tag in a.reviewers:
        p = os.path.join(a.returned, f"reviewer_{tag}.xlsx")
        if not os.path.isfile(p):
            sys.exit(f"reviewer {tag}: no returned workbook at {p}")
        sheets[tag] = _read_sheet(p)
        print(f"  reviewer {tag}: {len(sheets[tag])} labelled rows")

    common = set.intersection(*(set(s) for s in sheets.values())) if sheets else set()
    axis = lock.get("dispute_axis", DISPUTE_AXIS)
    disputed = sorted(u for u in common
                      if len({sheets[t][u].get(axis, "") for t in a.reviewers}) > 1)
    ordered = order_of(disputed, lock["seed"])

    out = {
        "schema_version": "reconciliation-order-v4",
        "lock_content_hash": lock["content_hash"],
        "rule_id": lock["rule_id"], "seed": lock["seed"], "dispute_axis": axis,
        "reviewers": list(a.reviewers),
        "n_common_rows": len(common),
        "n_disputed": len(disputed),
        "order": ordered,
        "derived_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "note": ("`order` is order_of(disputed, seed) and nothing else. "
                 "check_study_blinding_v4 recomputes it from the seed and refuses a mismatch, so "
                 "this file is a record rather than an authority."),
    }
    dest = os.path.join(STUDY_DIR, "RECONCILIATION_ORDER_V4.json")
    json.dump(out, open(dest, "w", encoding="utf-8"), indent=1, sort_keys=True)
    print(f"{len(disputed)} disputed of {len(common)} common rows on {axis}")
    print("wrote " + os.path.relpath(dest, SYS_ROOT))
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)
    rev = lambda x: tuple(t.strip() for t in x.split(",") if t.strip())

    lk = sub.add_parser("lock", help="commit the rule and seed before labelling")
    lk.add_argument("--reviewers", default="A,B,C", type=rev)
    lk.add_argument("--packages", default=os.path.join(C.ADJ_DIR, "SEND-HUMAN-V4"))
    lk.add_argument("--seed", type=int, default=SEED)
    lk.add_argument("--allow-labelled", action="store_true",
                    help="record a lock even though workbooks already carry labels. The lock then "
                         "states that it is not blind and the guard reports it.")
    lk.set_defaults(fn=cmd_lock)

    dv = sub.add_parser("derive", help="derive the order from the returned sheets")
    dv.add_argument("--reviewers", default="A,B,C", type=rev)
    dv.add_argument("--returned", default=os.path.join(C.ADJ_DIR, "RETURNED-V4"))
    dv.set_defaults(fn=cmd_derive)

    a = ap.parse_args()
    return a.fn(a)


if __name__ == "__main__":
    sys.exit(main())
