#!/usr/bin/env python3
"""Score the defect-level study on annotator B, and say plainly why not on both.

Two non-author annotators labelled the same stratified 200-finding sample from blinded packets.
Annotator B declared no knowledge of the study's objective; annotator A declared knowledge of it.
The protocol allows either, and requires the difference to be recorded, so B is the primary reading
and A is reported beside it as a sensitivity check rather than averaged into it.

The joint reconciliation of the 55 disputed findings is NOT used. That session resolved 41 of 55,
returned UNRELATED on all 41, adopted B on 41 of 41 and A on none, and its own working note shows
the rows were sorted by patch geometry computed outside the blinded packets: whether the finding's
file was patched, whether its function was, and the distance to the nearest changed line. Those are
the fields the packet builder withholds on purpose. The study exists to test whether patch geometry
overstates defect identification, so a label derived from patch geometry cannot be evidence about
patch geometry. The 55 stay reported as unresolved, which the protocol explicitly permits, and
dropping the reconciliation moves the pooled rate by 0.005.

Every rate here is a human judgment of whether a finding names the defect the vendor patched. The
geometric rates are recomputed on the same 200 findings so the two are read on one sample.

    python3 -m eval.defect_study_result_v3
"""
from __future__ import annotations
import json, os, sys, platform
from collections import Counter, defaultdict

import numpy as np

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
from eval import adjudication_v3_common as C
from eval.analyze_v3 import boot_rate, _slug_index

SYS_ROOT = C.SYS_ROOT
POP = os.path.join(SYS_ROOT, "revision-cns-v2", "data", "FINDING_POPULATION_V3.jsonl")
SAMPLE = os.path.join(SYS_ROOT, "revision-cns-v2", "out", "DEFECT_STUDY_SAMPLE_V3.json")
OUT = os.path.join(SYS_ROOT, "revision-cns-v2", "out", "DEFECT_STUDY_RESULT_V3.json")
REPS = 10000
SEED = 20260818
PRIMARY = "B"
AXES = C.TIER2_LABEL_AXES


def _kappa(a, b):
    """Cohen's kappa on two aligned label lists."""
    cats = sorted(set(a) | set(b))
    n = len(a)
    po = sum(1 for x, y in zip(a, b) if x == y) / n
    ca, cb = Counter(a), Counter(b)
    pe = sum((ca[c] / n) * (cb[c] / n) for c in cats)
    return (po - pe) / (1 - pe) if pe < 1 else None, po


def _boot_kappa(units, a_of, b_of, reps=REPS, seed=SEED):
    """Record-cluster bootstrap of kappa. Packets from one advisory are not independent, so the
    resampling unit is the advisory record, as the protocol specifies."""
    recs = sorted({u["record_uid"] for u in units})
    by = defaultdict(list)
    for u in units:
        by[u["record_uid"]].append(u)
    rng = np.random.default_rng(seed)
    out = []
    for _ in range(reps):
        pick = rng.integers(0, len(recs), size=len(recs))
        us = [u for i in pick for u in by[recs[i]]]
        k, _po = _kappa([a_of(u) for u in us], [b_of(u) for u in us])
        if k is not None:
            out.append(k)
    if not out:
        return [None, None]
    lo, hi = np.percentile(out, [2.5, 97.5])
    return [round(float(lo), 4), round(float(hi), 4)]


def boot_ratio(units, num, den, reps=REPS, seed=SEED):
    """Paired slug-cluster bootstrap of a ratio of two rates measured on the same findings.

    This is the ratio analogue of eval.analyze_v3.boot_diff and it is written to be the same
    procedure in every respect that makes an interval comparable. It indexes the clusters with the
    same eval.analyze_v3._slug_index, it draws one picks matrix of shape (reps, n_slugs) from
    numpy's default_rng at the same seed, and it reads the percentile interval at 2.5 and 97.5.
    boot_diff is the function it mirrors because boot_diff is the only existing bootstrap here that
    carries two statistics through one draw, which is what pairing means: the same resampled
    findings feed the numerator and the denominator, so the correlation between them is inside the
    replicate rather than assumed away.

    The interval this replaces divided a fixed point estimate of the numerator by the endpoints of
    the denominator's own interval. That is wrong twice over. It gives the numerator no variance at
    all, and it treats two rates read off one sample of 200 findings as if the sample had been drawn
    twice. Both quantities are recomputed here inside each replicate.

    num(u) and den(u) return booleans. The ratio is undefined in a replicate that draws no
    denominator event, so those replicates are dropped and counted rather than being given a value.
    """
    slugs, idx = _slug_index(units)
    kn = np.zeros(len(slugs))
    kd = np.zeros(len(slugs))
    n = np.zeros(len(slugs))
    for u in units:
        i = idx[u["slug"]]
        n[i] += 1
        kn[i] += 1 if num(u) else 0
        kd[i] += 1 if den(u) else 0
    if not kd.sum():
        return {"point": None, "ci95": [None, None], "replicates": reps,
                "replicates_defined": 0, "seed": seed, "cluster_unit": "plugin_slug"}
    # both rates share the denominator n, so the ratio of rates is the ratio of counts
    point = kn.sum() / kd.sum()
    rng = np.random.default_rng(seed)
    picks = rng.integers(0, len(slugs), size=(reps, len(slugs)))
    Kn = kn[picks].sum(1)
    Kd = kd[picks].sum(1)
    ok = Kd > 0
    ratios = np.divide(Kn, Kd, out=np.full(reps, np.nan), where=ok)
    lo, hi = np.nanpercentile(ratios, [2.5, 97.5])
    return {"point": round(float(point), 2),
            "ci95": [round(float(lo), 2), round(float(hi), 2)],
            "replicates": reps, "replicates_defined": int(ok.sum()), "seed": seed,
            "cluster_unit": "plugin_slug",
            "share_at_or_below_one": round(float(np.mean(ratios[ok] <= 1.0)), 4)}


SHIPPED = os.path.join(ROOT, "defect-study", "defect_study_labels.csv")


def _frame_composition():
    """Per-tool composition of the population the 200 findings were drawn from.

    A reviewer counted the sample's 11 Progpilot findings against the corpus finding population,
    where Progpilot is 874 of 8467 and 21 of 200 would be proportional, and read the sample as short
    of Progpilot. The frame is this file, where Progpilot's share is less than half of that, so 11
    is above proportional and not below it. Neither count was wrong. The paper never named which
    population the sample came from, so this records it.
    """
    if not os.path.isfile(POP):
        return Counter()
    with open(POP, encoding="utf-8") as fh:
        return Counter(json.loads(ln)["tool"] for ln in fh if ln.strip())


def _units_from_shipped():
    """Rebuild the analysis units from the shipped anonymised label sheet.

    A reviewer objected that the one number in this paper resting on human judgment could not be
    recomputed from the bundle, because the sheets lived outside it. They ship now, and this is the
    path both the bundle and the working tree take when the sheet is present, so the published
    result is checked by exactly the file a reader holds.
    """
    import csv as _csv
    out = []
    with open(SHIPPED, encoding="utf-8") as fh:
        for r in _csv.DictReader(fh):
            u = {k: r[k] for k in ("packet_id", "record_uid", "slug", "cve", "tool",
                                   "advisory_class")}
            for g in ("in_patched_file", "same_callable_as_change", "on_exact_changed_line"):
                u[g] = str(r[g]).strip().lower() == "true"
            for who in ("A", "B"):
                u[who] = {ax: r[f"{who}_{ax}"] for ax in AXES}
            out.append(u)
    return out


def main():
    if os.path.isfile(SHIPPED):
        return _score(_units_from_shipped(), _frame_composition())
    key = C.read_json(os.path.join(C.TIER2_DIR, "BLINDING_KEY.json"))["payload"]["map"]
    labels = {r: C.read_json(os.path.join(C.TIER2_DIR, f"reviewer_{r}_findings.json"))["payload"]["labels"]
              for r in ("A", "B")}

    smp = C.read_json(SAMPLE)
    smp = smp.get("payload", smp)
    want = set(smp["sample"]["finding_uids"])

    geo = {}
    with open(POP, encoding="utf-8") as fh:
        for ln in fh:
            if ln.strip():
                r = json.loads(ln)
                geo[r["finding_uid"]] = r

    units = []
    for pid, e in key.items():
        if e["finding_uid"] not in want:
            continue
        g = geo[e["finding_uid"]]
        units.append({"packet_id": pid, "finding_uid": e["finding_uid"],
                      "record_uid": e["record_uid"], "slug": g["slug"], "cve": g["cve"],
                      "tool": e["tool"], "advisory_class": g["advisory_class"],
                      "in_patched_file": bool(g["in_patched_file"]),
                      "same_callable_as_change": bool(g["same_callable_as_change"]),
                      "on_exact_changed_line": bool(g["on_exact_changed_line"]),
                      "A": labels["A"][pid], "B": labels["B"][pid]})
    if len(units) != len(want):
        raise SystemExit(f"sample has {len(want)} findings but only {len(units)} joined to a packet")
    return _score(units, _frame_composition())


def _score(units, frame):
    _meta_p = os.path.join(C.TIER1_DIR, "REVIEWER_METADATA_TEMPLATE.json")
    if not os.path.isfile(_meta_p):
        _meta_p = os.path.join(ROOT, "defect-study", "ANNOTATOR_METADATA.json")
    meta = C.read_json(_meta_p)["payload"]

    def same_defect(who):
        return lambda u: u[who]["root_cause_relation"] == "SAME_DEFECT"

    res = {"per_tool": {}, "pooled": {}, "geometry_same_sample": {}}
    tools = sorted({u["tool"] for u in units})
    for who in ("A", "B"):
        res["pooled"][who] = boot_rate(units, same_defect(who), REPS, SEED)
        for t in tools:
            sub = [u for u in units if u["tool"] == t]
            res["per_tool"].setdefault(t, {})[who] = boot_rate(sub, same_defect(who), REPS, SEED)

    # the same 200 findings, scored geometrically, so the contrast is read on one sample
    for name, f in (("in_patched_file", lambda u: u["in_patched_file"]),
                    ("same_callable_as_change", lambda u: u["same_callable_as_change"]),
                    ("on_exact_changed_line", lambda u: u["on_exact_changed_line"])):
        res["geometry_same_sample"][name] = {"pooled": boot_rate(units, f, REPS, SEED)}
        for t in tools:
            sub = [u for u in units if u["tool"] == t]
            res["geometry_same_sample"][name][t] = boot_rate(sub, f, REPS, SEED)

    # agreement between the two readings, on every axis, with a record-clustered interval
    res["agreement"] = {}
    for ax in AXES:
        a = [u["A"][ax] for u in units]
        b = [u["B"][ax] for u in units]
        k, po = _kappa(a, b)
        res["agreement"][ax] = {
            "n": len(units), "observed_agreement": round(po, 4),
            "cohens_kappa": None if k is None else round(k, 4),
            "kappa_ci95_record_cluster": _boot_kappa(units, lambda u, ax=ax: u["A"][ax],
                                                     lambda u, ax=ax: u["B"][ax])}

    # The excluded reconciliation, measured rather than described. Its numbers appear in the
    # supplement, so they have to be generated like every other number in this paper, and the
    # archived workbook is the source.
    import openpyxl
    # The archived workbook lives outside the bundle, so inside the bundle this block used to come
    # out None while the working tree filled it. That is a real reproducibility hole and the
    # reproduce kit caught it as a MISMATCH: the supplement prints n_resolved, adopted_annotator_B
    # and pooled_rate_if_included from here, and a reader could not recompute any of them. The
    # workbook ships now, and the shipped copy is preferred, so both trees take the same path.
    RECON_XLSX = os.path.join(ROOT, "defect-study", "reconciliation_returned_excluded.xlsx")
    if not os.path.isfile(RECON_XLSX):
        RECON_XLSX = os.path.join(C.ADJ_DIR, "RETURNED-2026-08-17",
                                  "reconciliation_returned_excluded.xlsx")
    excl = None
    if os.path.isfile(RECON_XLSX):
        ws = openpyxl.load_workbook(RECON_XLSX)["DOI CHIEU"]
        hdr = [c.value for c in ws[1]]
        rws = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
        fin = {r["packet_id"]: str(r.get("final_root_cause_relation") or "").strip() for r in rws}
        res_rows = {k: v for k, v in fin.items() if v}
        vals = Counter(res_rows.values())
        to_a = sum(1 for r in rws if r["packet_id"] in res_rows
                   and res_rows[r["packet_id"]] == str(r.get("A_root_cause")))
        to_b = sum(1 for r in rws if r["packet_id"] in res_rows
                   and res_rows[r["packet_id"]] == str(r.get("B_root_cause")))
        by_pid = {u["packet_id"]: u for u in units}
        def merged(u):
            if u["A"]["root_cause_relation"] == u["B"]["root_cause_relation"]:
                return u["A"]["root_cause_relation"]
            return res_rows.get(u["packet_id"])
        inc = [u for u in units if merged(u) is not None]
        n_sd = sum(1 for u in units if merged(u) == "SAME_DEFECT")
        excl = {"n_resolved": len(res_rows), "n_disputed": len(rws),
                "value_spread": dict(vals),
                "adopted_annotator_A": to_a, "adopted_annotator_B": to_b,
                "pooled_rate_if_included": round(n_sd / len(units), 4),
                "source": os.path.basename(RECON_XLSX),
                "excluded_because": ("its working note shows the rows were pre-sorted by patch "
                                     "geometry computed outside the blinded packets, and a defect "
                                     "label derived from the geometry under test cannot be "
                                     "evidence about that geometry")}
    res["excluded_reconciliation"] = excl

    disputed = [u for u in units if u["A"]["root_cause_relation"] != u["B"]["root_cause_relation"]]
    res["disagreement"] = {
        "n_disputed_root_cause": len(disputed),
        "n_total": len(units),
        "resolution": "left unresolved",
        "why": ("The joint reconciliation session was excluded. It resolved 41 of 55, returned "
                "UNRELATED on all 41, adopted annotator B on 41 of 41 and annotator A on none, and "
                "its working note shows the rows were sorted by patch geometry computed outside the "
                "blinded packets. A defect label derived from the geometry the study is testing "
                "cannot serve as evidence about that geometry. Excluding it moves the pooled rate "
                "by 0.005.")}

    # The overstatement factor, which is what the paper's claim actually is, with the interval that
    # comes from the expert rate rather than a point ratio. A reviewer noted that an earlier draft
    # called choosing the blind reading conservative. It is not: the blind rate is the LOWER of the
    # two, so it yields the LARGER factor and the stronger claim. The design reason for choosing it
    # stands on its own and does not need a false modesty argument, so the factor is computed for
    # both readings and both are printed.
    gf = res["geometry_same_sample"]["in_patched_file"]["pooled"]["rate"]
    res["overstatement_factor"] = {"geometric_rate": gf}
    for who in ("A", "B"):
        pl = res["pooled"][who]
        lo, hi = pl["ci95"]
        res["overstatement_factor"][who] = {
            "point": round(gf / pl["rate"], 2) if pl["rate"] else None,
            "from_rate_ci95": [round(gf / hi, 1) if hi else None,
                               round(gf / lo, 1) if lo else None],
            "paired_cluster_bootstrap_ci95": boot_ratio(
                units, lambda u: u["in_patched_file"], same_defect(who))}
    res["overstatement_factor"]["note"] = (
        "The blind reading gives the larger factor, so reporting it is not the conservative choice. "
        "It is chosen because it is the blind one, and the interval is carried so the factor is not "
        "read as a constant. Two intervals are carried for that factor and they are not "
        "interchangeable. from_rate_ci95 is the superseded one: it divides the geometric rate, held "
        "at its point estimate, by the endpoints of the annotator's own rate interval, so it gives "
        "the numerator no uncertainty and it treats two rates measured on one sample of 200 "
        "findings as if the sample had been drawn twice. It is kept because the earlier prose "
        "quoted it and deleting it would erase that record. paired_cluster_bootstrap_ci95 is the "
        "one the paper should cite: it resamples plugin slugs with replacement, the same cluster "
        "and the same seed every other interval in this paper uses, recomputes both the geometric "
        "rate and the annotator's same-defect rate inside each replicate, and reads the percentile "
        "interval of the ratio, so the pairing between numerator and denominator is preserved.")

    res["annotators"] = {
        w: {k: meta[f"reviewer_{w}"].get(k) for k in
            ("reviewer_pseudonym", "years_experience", "knows_research_objective",
             "is_paper_author", "conflict_of_interest")}
        for w in ("A", "B")}
    res["primary_reading"] = {
        "annotator": PRIMARY,
        "reason": ("Annotator B declared no knowledge of the study's objective while labelling and "
                   "annotator A declared knowledge of it. Both declared they are not authors. B is "
                   "therefore the blind reading and is reported as primary; A is reported beside it "
                   "and is the more generous of the two, so the primary rate is the lower one.")}
    _fn = int(sum(frame.values()))
    if _fn:
        res["frame_composition"] = {
            "population": os.path.relpath(POP, SYS_ROOT),
            "n_findings": _fn,
            "per_tool": {t: {"n_frame": int(frame[t]),
                             "share": round(frame[t] / _fn, 4),
                             "expected_in_sample": round(len(units) * frame[t] / _fn, 1),
                             "n_sample": sum(1 for u in units if u["tool"] == t)}
                         for t in sorted(frame)},
            "note": ("the sample is drawn from this population, so a tool's share here is what "
                     "its count in the sample is to be read against, not its share of the corpus "
                     "finding population"),
        }
    res["config"] = {"sample": os.path.relpath(SAMPLE, SYS_ROOT), "n_findings": len(units),
                     "n_records": len({u["record_uid"] for u in units}),
                     "n_slugs": len({u["slug"] for u in units}),
                     "bootstrap_replicates": REPS, "seed": SEED,
                     "rate_bootstrap_unit": "plugin_slug",
                     "ratio_bootstrap_unit": "plugin_slug",
                     "kappa_bootstrap_unit": "advisory_record",
                     "python": platform.python_version()}
    C.write_json(OUT, C.envelope("defect_study_result", res))

    p = res["pooled"]
    print(f"pooled same_defect   B(primary) {p['B']['rate']} CI{p['B']['ci95']}"
          f"   A {p['A']['rate']} CI{p['A']['ci95']}   n={p['B']['n']}")
    g = res["geometry_same_sample"]
    print(f"same sample, geometry: in_patched_file {g['in_patched_file']['pooled']['rate']}"
          f"  same_callable {g['same_callable_as_change']['pooled']['rate']}"
          f"  exact_line {g['on_exact_changed_line']['pooled']['rate']}")
    for t in tools:
        print(f"  {t:10} B {res['per_tool'][t]['B']['rate']} CI{res['per_tool'][t]['B']['ci95']}"
              f"   file {g['in_patched_file'][t]['rate']}   n={res['per_tool'][t]['B']['n']}")
    of = res["overstatement_factor"]
    for who in ("A", "B"):
        pc = of[who]["paired_cluster_bootstrap_ci95"]
        print(f"overstatement {who}: point {of[who]['point']}"
              f"   paired slug-cluster CI {pc['ci95']}"
              f"   superseded plug-in CI {of[who]['from_rate_ci95']}"
              f"   reps {pc['replicates_defined']}/{pc['replicates']} seed {pc['seed']}")
    print(f"root-cause disagreements left unresolved: {len(disputed)}/{len(units)}")
    print("wrote", os.path.relpath(OUT, SYS_ROOT))
    return 0


if __name__ == "__main__":
    sys.exit(main())
