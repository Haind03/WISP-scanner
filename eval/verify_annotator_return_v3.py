#!/usr/bin/env python3
"""Decide whether a returned annotator workbook is an independent human reading.

A returned sheet is not evidence just because a human sent it back. The protocol asks for five
SEPARATE judgments per finding and a written reason for each, so the checks here look for the
signature of a sheet that was filled by a program and then signed off, rather than judged.

Seven tests, each of which a genuine reading passes and a generated sheet fails:

  T1 structure       identifier and context columns byte-identical to what was sent, no row added,
                     removed, reordered or duplicated
  T2 completeness    the judgment columns are filled
  T3 abstention      the abstention values exist in the returned data, because a reading with no
                     abstention anywhere is a reading that never met a hard case
  T4 axis-independence
                     the three secondary axes are not a function of the two primary ones. This is
                     the load-bearing test. The protocol exists to collect five separate judgments,
                     so if evidence_quality, confidence and reason_code are each single-valued
                     inside every (class_relation, root_cause_relation) cell, the sheet carries two
                     judgments wearing five hats
  T5 reason text     the written reasons are individual, not a template with slots, and do not all
                     end in the same sentence
  T6 label/text tie  the written reason does not map one-to-one onto the label it accompanies. A
                     reason that is a rendering of the label justifies nothing
  T7 independence    the labels do not agree with an archived machine draft more than with the
                     other human annotators
  T8 tier 1 prose    the defect cards are written per advisory, not drawn from a handful of canned
                     sentences, and patch_mechanism describes a mechanism rather than pasting code

Usage:
    python3 -m eval.verify_annotator_return_v3 <returned.xlsx> [--sent <sent.xlsx>] [--peer <x.xlsx> ...]
    python3 -m eval.verify_annotator_return_v3 --all
"""
from __future__ import annotations
import os, sys, difflib, statistics, argparse
from collections import Counter, defaultdict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
from eval import adjudication_v3_common as C

AXES = C.TIER2_LABEL_AXES
PRIMARY_AXES = ["class_relation", "root_cause_relation"]
SECONDARY_AXES = [a for a in AXES if a not in PRIMARY_AXES]
FROZEN_T2 = ["packet_id", "advisory_class", "finding_file", "finding_line", "packet_path"]
ABSTENTIONS = {"root_cause_relation": "INSUFFICIENT_EVIDENCE",
               "evidence_quality": "INSUFFICIENT",
               "class_relation": "UNCERTAIN"}


def _tier2(path):
    ws = openpyxl.load_workbook(path)["Tier2"]
    hdr = [c.value for c in ws[1]]
    rows = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return hdr, {r["packet_id"]: r for r in rows}, rows


class Report:
    def __init__(self, name):
        self.name, self.rows, self.fails = name, [], 0

    def add(self, tid, title, ok, detail):
        self.rows.append((tid, title, ok, detail))
        if not ok:
            self.fails += 1

    def show(self):
        print(f"\n=== {self.name} ===")
        for tid, title, ok, detail in self.rows:
            print(f"  [{'ok' if ok else 'FAIL'}] {tid} {title}")
            for ln in detail.splitlines():
                print(f"          {ln}")
        print(f"  verdict: {'ACCEPT' if not self.fails else 'REJECT'}"
              f"  ({self.fails} of {len(self.rows)} tests failed)")
        return self.fails


def verify(returned, sent=None, peers=(), drafts=()):
    rep = Report(os.path.basename(returned))
    hdr, R, rows = _tier2(returned)

    # T1 structure
    if sent:
        shdr, S, _ = _tier2(sent)
        bad = []
        if shdr != hdr:
            bad.append("header changed")
        if set(S) != set(R):
            bad.append(f"packet id set changed ({len(S)} sent, {len(R)} back)")
        if len(rows) != len(R):
            bad.append("duplicate packet_id in the return")
        for col in FROZEN_T2:
            if col in hdr and any(S[p][col] != R[p][col] for p in S if p in R):
                bad.append(f"frozen column edited: {col}")
        rep.add("T1", "structure unchanged from the sent workbook",
                not bad, "\n".join(bad) or f"{len(R)} rows, {len(FROZEN_T2)} frozen columns intact")
    else:
        rep.add("T1", "structure unchanged from the sent workbook", True, "no sent workbook given, skipped")

    # T2 completeness
    blank = {a: sum(1 for r in rows if not str(r.get(a) or "").strip()) for a in AXES}
    rep.add("T2", "judgment columns filled", all(v < len(rows) for v in blank.values()),
            "blank per axis: " + ", ".join(f"{k} {v}" for k, v in blank.items()))

    # T3 abstention used somewhere
    used = {a: sum(1 for r in rows if r.get(a) == v) for a, v in ABSTENTIONS.items()}
    rep.add("T3", "abstention values appear in the reading", sum(used.values()) > 0,
            "counts: " + ", ".join(f"{a}={n}" for a, n in used.items()))

    # T4 axis independence, the load-bearing one
    cells = defaultdict(lambda: defaultdict(set))
    for r in rows:
        k = tuple(r[a] for a in PRIMARY_AXES)
        for a in SECONDARY_AXES:
            cells[k][a].add(r[a])
    det = {a: sum(1 for k in cells if len(cells[k][a]) == 1) for a in SECONDARY_AXES}
    n_cells = len(cells)
    collapsed = [a for a in SECONDARY_AXES if det[a] >= n_cells - 1]
    rep.add("T4", "secondary axes carry judgment of their own", not collapsed,
            f"{n_cells} (class_relation, root_cause_relation) cells; single-valued in "
            + ", ".join(f"{a} {det[a]}/{n_cells}" for a in SECONDARY_AXES)
            + ("\ncollapsed axes: " + ", ".join(collapsed) if collapsed else ""))

    # T5 reason text individual
    notes = [str(r.get("notes") or "").strip() for r in rows]
    nonempty = [n for n in notes if n]
    distinct = len(set(nonempty))
    tails = Counter(n.rsplit(".", 2)[-2].strip() + "." if n.count(".") >= 2 else n for n in nonempty)
    top_tail, top_n = tails.most_common(1)[0] if tails else ("", 0)
    ratio = distinct / len(nonempty) if nonempty else 0.0
    ok5 = ratio >= 0.5 and top_n < 0.5 * len(nonempty)
    rep.add("T5", "written reasons are individual", ok5,
            f"{len(nonempty)} reasons, {distinct} distinct ({ratio:.2f})\n"
            f"most repeated closing sentence appears {top_n} times: {top_tail[:90]!r}")

    # T6 the reason is not a rendering of the label
    worst = None
    for a in PRIMARY_AXES:
        buckets = defaultdict(set)
        for r, n in zip(rows, notes):
            # bucket a note by its distinct sentence set, so a template with slots collapses
            sig = tuple(sorted({s.strip()[:40] for s in n.split(".") if s.strip()}))
            buckets[sig].add(r[a])
        pure = sum(1 for v in buckets.values() if len(v) == 1)
        # how well does the note signature predict the label
        cov = pure / len(buckets) if buckets else 0
        if worst is None or cov > worst[1]:
            worst = (a, cov, len(buckets))
    ok6 = not (worst and worst[1] >= 0.99 and worst[2] < len(rows) * 0.6)
    rep.add("T6", "reason text is not a one-to-one rendering of the label", ok6,
            f"note signatures predict {worst[0]} with purity {worst[1]:.2f} over {worst[2]} "
            f"signatures for {len(rows)} rows")

    # T7 independence from an archived machine draft
    if drafts and peers:
        def agree(D):
            common = [p for p in R if p in D]
            return {a: sum(1 for p in common if R[p][a] == D[p][a]) / len(common) for a in AXES}, len(common)
        lines, bad7 = [], []
        peer_ag = {}
        for pth in peers:
            _h, P, _r = _tier2(pth)
            peer_ag[os.path.basename(pth)], n = agree(P)
            lines.append(f"vs human {os.path.basename(pth)} (n={n}): "
                         + ", ".join(f"{a} {peer_ag[os.path.basename(pth)][a]:.2f}" for a in AXES))
        for pth in drafts:
            _h, D, _r = _tier2(pth)
            dag, n = agree(D)
            lines.append(f"vs draft {os.path.basename(pth)} (n={n}): "
                         + ", ".join(f"{a} {dag[a]:.2f}" for a in AXES))
            for a in AXES:
                best_h = max(v[a] for v in peer_ag.values())
                if dag[a] > best_h + 0.15:
                    bad7.append(f"{a}: agrees with {os.path.basename(pth)} at {dag[a]:.2f} but with "
                                f"the closest human at only {best_h:.2f}")
        rep.add("T7", "labels not closer to a machine draft than to the human annotators",
                not bad7, "\n".join(lines + bad7))
    else:
        rep.add("T7", "labels not closer to a machine draft than to the human annotators", True,
                "no draft or peer given, skipped")

    # T8 tier 1 defect cards written per advisory
    wb = openpyxl.load_workbook(returned)
    if "Tier1" in wb.sheetnames:
        ws = wb["Tier1"]
        h1 = [c.value for c in ws[1]]
        t1 = [dict(zip(h1, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
        summ = [str(r.get("root_cause_summary") or "").strip() for r in t1]
        reused = sum(n for _s, n in Counter([x for x in summ if x]).items() if n > 1)
        ann = [str(r.get("annotation") or "").strip() for r in t1]
        stamp, stamp_n = Counter([a for a in ann if a]).most_common(1)[0] if any(ann) else ("", 0)
        pasted = sum(1 for r in t1 if str(r.get("patch_mechanism") or "").count(" | ") >= 2)
        bad8 = []
        if reused:
            bad8.append(f"{reused} of {len(t1)} defect cards reuse a canned root_cause_summary")
        if stamp_n > 0.1 * len(t1):
            bad8.append(f"{stamp_n} of {len(t1)} annotations are the identical sentence "
                        f"{stamp[:80]!r}")
        if pasted > 0.1 * len(t1):
            bad8.append(f"{pasted} of {len(t1)} patch_mechanism cells are pasted code lines "
                        f"rather than a described mechanism")
        rep.add("T8", "defect cards written per advisory", not bad8,
                "\n".join(bad8) or f"{len(t1)} cards, {len(set(summ))} distinct summaries, "
                f"{pasted} pasted patch_mechanism cells")
    else:
        rep.add("T8", "defect cards written per advisory", True, "no Tier1 sheet, skipped")
    return rep


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("returned", nargs="?")
    ap.add_argument("--sent")
    ap.add_argument("--peer", action="append", default=[])
    ap.add_argument("--draft", action="append", default=[])
    ap.add_argument("--all", action="store_true")
    a = ap.parse_args()

    ret17 = os.path.join(C.ADJ_DIR, "RETURNED-2026-08-17")
    rejc = os.path.join(C.ADJ_DIR, "REJECTED-C-2026-08-18")
    drafts = [os.path.join(C.ADJ_DIR, "AI-DRAFT-2026-08-03", f) for f in
              ("adjudication_reviewer_A_AI_DRAFT.xlsx", "adjudication_reviewer_B_FILLED_1824.xlsx")]
    drafts = [d for d in drafts if os.path.isfile(d)]

    if a.all:
        jobs = [
            (os.path.join(ret17, "reviewer_A.xlsx"),
             os.path.join(C.ADJ_DIR, "SEND-HUMAN-2026-08-17", "reviewer_A", "reviewer_A.xlsx"),
             [os.path.join(ret17, "reviewer_B.xlsx")]),
            (os.path.join(ret17, "reviewer_B.xlsx"), None,
             [os.path.join(ret17, "reviewer_A.xlsx")]),
            (os.path.join(rejc, "reviewer_C_reviewed.xlsx"),
             os.path.join(C.ADJ_DIR, "SEND-HUMAN-C-2026-08-18", "reviewer_C", "reviewer_C.xlsx"),
             [os.path.join(ret17, "reviewer_A.xlsx"), os.path.join(ret17, "reviewer_B.xlsx")]),
        ]
        fails = 0
        for ret, sent, peers in jobs:
            if not os.path.isfile(ret):
                print(f"\n=== {os.path.basename(ret)} === not present, skipped")
                continue
            fails += verify(ret, sent if sent and os.path.isfile(sent) else None,
                            [p for p in peers if os.path.isfile(p)], drafts).show()
        print(f"\ntotal failing tests across all workbooks: {fails}")
        return 0
    if not a.returned:
        ap.error("give a workbook or --all")
    verify(a.returned, a.sent, a.peer, a.draft or drafts).show()
    return 0


if __name__ == "__main__":
    sys.exit(main())
