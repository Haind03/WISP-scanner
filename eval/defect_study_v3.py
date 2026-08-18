#!/usr/bin/env python3
"""Draw the defect-level study sample under the locked protocol, and build the two workbooks.

The protocol in `revision-cns-v2/bundle-src/DEFECT-LEVEL-STUDY-PROTOCOL.md` was fixed before any
labelling, and it specifies a stratified sample of the finding population balanced across
**vulnerability class, patch shape and plugin size**, roughly 180 to 220 findings, drawn
proportionally with a floor per stratum so every class and every patch shape is represented, with
the sample, its seed and its per-stratum counts shipped alongside.

The older `eval/build_adjudication_v3.py` stratifies on (class x rank) over two tools. That is a
different design, and running it here would produce a study that does not match the protocol the
annotators were shown. This script implements the protocol instead.

What it does not do is re-derive evidence. The 710 blinded packets and the 100 defect-card contexts
for the matched sample already exist and were built from the archives. This selects a subset of
them, joins through the sealed key by `finding_uid`, and never opens which tool a packet came from
for any purpose other than the diagnostic count it prints, which is written to the sample file so
the author can see the design's consequence rather than discover it after labelling.

    python3 -m eval.defect_study_v3 --target 200          # sample + package + workbooks
    python3 -m eval.defect_study_v3 --sample-only
"""
from __future__ import annotations
import argparse, csv, hashlib, json, os, platform, random, shutil, sys, time, zipfile
from collections import Counter, defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
from eval import adjudication_v3_common as C
from eval import reviewer_xlsx_v3 as X

from openpyxl import Workbook

SYS_ROOT = C.SYS_ROOT
OUT_DIR = os.path.join(SYS_ROOT, "revision-cns-v2", "out")
POP = os.path.join(SYS_ROOT, "revision-cns-v2", "data", "FINDING_POPULATION_V3.jsonl")
SHAPES = os.path.join(OUT_DIR, "PATCH_SHAPE_CENSUS_V3.json")
ARCHIVES = os.path.join(SYS_ROOT, "patchstack_bugbounty", "plugins_01")
SAMPLE_OUT = os.path.join(OUT_DIR, "DEFECT_STUDY_SAMPLE_V3.json")

TOPK = 3
SEED = 20260810
TARGET_MIN, TARGET_MAX = 180, 220
REFACTOR_FILES = 20          # a release touching this many PHP files is a refactor, not a fix
SIZE_LABELS = ("small", "medium", "large")


# ------------------------------------------------------------------ stratum axes


def patch_shape(rec):
    """The protocol's five shapes, in priority order, from the shipped patch-shape census.

    Priority matters: a release that renames files and also inserts a guard is a refactor-heavy
    release for sampling purposes, because that is what makes it hard to localize.

    `guard-insertion` is a FILE-level test, at least one changed PHP file only gained lines. The
    record-level test, that the release changed no vulnerable-side line anywhere, matches nothing:
    every one of the hundred records touches at least one existing line, because a real release
    also bumps a version constant or a changelog. Reading the strict test as "this corpus contains
    no guard insertions" would be wrong about the corpus, so the looser proxy is used and named.
    """
    if rec.get("n_php_scored_files", 0) >= REFACTOR_FILES:
        return "refactor-heavy"
    if rec.get("n_php_renamed", 0) > 0:
        return "rename"
    if rec.get("n_php_deleted", 0) > 0:
        return "deletion"
    if rec.get("n_php_pure_insertion", 0) > 0:
        return "guard-insertion"
    return "code-replacement"


def php_file_count(slug):
    """Plugin size as the number of PHP files in the vulnerable archive.

    Read from the zip's central directory, so nothing is extracted. A slug whose archive is not on
    this machine returns None and lands in its own size bucket rather than being guessed at.
    """
    d = os.path.join(ARCHIVES, slug)
    if not os.path.isdir(d):
        return None
    cands = sorted(f for f in os.listdir(d) if f.upper().endswith("-VULNERABLE.ZIP"))
    if not cands:
        return None
    try:
        with zipfile.ZipFile(os.path.join(d, cands[0])) as z:
            return sum(1 for n in z.namelist() if n.lower().endswith(".php"))
    except (zipfile.BadZipFile, OSError):
        return None


def size_bucket(n, cuts):
    if n is None:
        return "unknown"
    return SIZE_LABELS[0] if n <= cuts[0] else (SIZE_LABELS[1] if n <= cuts[1] else SIZE_LABELS[2])


# ------------------------------------------------------------------ allocation


def allocate(sizes, target):
    """Proportional allocation with a floor of one per non-empty stratum, largest remainder.

    The floor is what the protocol asks for and it is not free: with many thin strata the floors
    alone can exceed the target, so the function reports that rather than silently overshooting.
    """
    strata = [s for s, n in sizes.items() if n > 0]
    if len(strata) > target:
        raise SystemExit(f"{len(strata)} non-empty strata but target {target}: raise the target "
                         f"or coarsen a stratum axis, do not drop the floor")
    total = sum(sizes[s] for s in strata)
    alloc = {s: 1 for s in strata}
    left = target - len(strata)
    if left > 0:
        share = {s: (sizes[s] - 1) / max(1, total - len(strata)) * left for s in strata}
        base = {s: int(share[s]) for s in strata}
        rem = sorted(strata, key=lambda s: (-(share[s] - base[s]), s))
        give = left - sum(base.values())
        for s in rem[:give]:
            base[s] += 1
        for s in strata:
            alloc[s] += base[s]
    # never ask a stratum for more findings than it holds, and hand the surplus to the largest
    for s in strata:
        if alloc[s] > sizes[s]:
            surplus = alloc[s] - sizes[s]
            alloc[s] = sizes[s]
            for t in sorted(strata, key=lambda t: -(sizes[t] - alloc[t])):
                room = sizes[t] - alloc[t]
                if room <= 0:
                    continue
                take = min(room, surplus)
                alloc[t] += take
                surplus -= take
                if not surplus:
                    break
    return alloc


# ------------------------------------------------------------------ the sample


def draw(target):
    units = [json.loads(l) for l in open(POP, encoding="utf-8") if l.strip()]
    units = [u for u in units if u["rank"] <= TOPK]

    census = json.load(open(SHAPES, encoding="utf-8"))["datasets"]["matched-100"]["records"]
    by_key = {r["key"]: r for r in census}

    slugs = sorted({u["slug"] for u in units})
    sizes = {s: php_file_count(s) for s in slugs}
    known = sorted(v for v in sizes.values() if v is not None)
    if len(known) < 3:
        raise SystemExit("plugin archives are not on this machine, cannot size-stratify")
    cuts = (known[len(known) // 3], known[2 * len(known) // 3])

    for u in units:
        rec = by_key.get(u["slug"] + "|" + u["cve"])
        if rec is None:
            raise SystemExit("no patch-shape row for " + u["slug"] + "|" + u["cve"])
        u["_shape"] = patch_shape(rec)
        u["_size"] = size_bucket(sizes[u["slug"]], cuts)
        u["_stratum"] = "%s|%s|%s" % (u["advisory_class"], u["_shape"], u["_size"])

    pop = defaultdict(list)
    for u in units:
        pop[u["_stratum"]].append(u)
    counts = {s: len(v) for s, v in pop.items()}
    alloc = allocate(counts, target)

    rng = random.Random(SEED)
    picked = []
    for s in sorted(pop):
        cand = sorted(pop[s], key=lambda u: u["finding_uid"])
        picked.extend(rng.sample(cand, alloc[s]))
    picked.sort(key=lambda u: u["finding_uid"])

    strata_table = {s: {"population": counts[s], "allocated": alloc[s],
                        "inclusion_probability": round(alloc[s] / counts[s], 6)}
                    for s in sorted(counts)}
    return units, picked, strata_table, cuts, sizes


# ------------------------------------------------------------------ package


def _rows_for(picked, key_map, pkt_dir_rel):
    """Tier-2 CSV rows for the sampled findings, joined to their packets through the sealed key."""
    by_uid = {}
    for pid, e in key_map.items():
        by_uid[e["finding_uid"]] = pid
    packets = {p["packet_id"]: p for p in
               C.read_json(os.path.join(C.TIER2_DIR, "PACKETS.json"))["payload"]["packets"]}
    rows, missing = [], []
    for u in picked:
        pid = by_uid.get(u["finding_uid"])
        if pid is None or pid not in packets:
            missing.append(u["finding_uid"])
            continue
        p = packets[pid]
        rows.append({"packet_id": pid, "advisory_class": p["advisory_class"],
                     "finding_file": p["finding_file"], "finding_line": p["finding_line"],
                     "packet_file": "%s/%s.md" % (pkt_dir_rel, pid),
                     "class_relation": "", "root_cause_relation": "", "evidence_quality": "",
                     "confidence": "", "reason_code": "", "notes": ""})
    if missing:
        raise SystemExit("%d sampled finding(s) have no packet, refusing to build a partial study"
                         % len(missing))
    rows.sort(key=lambda r: r["packet_id"])
    return rows


def _tier1_rows(picked, key_map):
    want = {key_map[pid]["record_uid"]: key_map[pid]
            for pid, e in key_map.items()
            if e["finding_uid"] in {u["finding_uid"] for u in picked}}
    ctx = C.read_json(os.path.join(C.TIER1_DIR, "DEFECT_CARDS_CONTEXT.json"))["payload"]["cards"]
    by_uid = {c["record_uid"]: c for c in ctx}
    rows = []
    for ruid in sorted(want):
        c = by_uid[ruid]
        rows.append({"record_uid": ruid, "slug": c["slug"], "cve": c["cve"],
                     "advisory_class": want[ruid]["advisory_class"],
                     "context_file": "context/%s.md" % ruid,
                     **{k: "" for k in X.T1_FILL}})
    return rows


def build_package(picked, dest, tag):
    key = C.read_json(os.path.join(C.TIER2_DIR, "BLINDING_KEY.json"))["payload"]["map"]
    t2 = _rows_for(picked, key, "packets")
    t1 = _tier1_rows(picked, key)

    pkg = os.path.join(dest, "reviewer_%s" % tag)
    shutil.rmtree(pkg, ignore_errors=True)
    os.makedirs(os.path.join(pkg, "tier1", "context"), exist_ok=True)
    os.makedirs(os.path.join(pkg, "tier2", "packets"), exist_ok=True)
    for r in t1:
        shutil.copy2(os.path.join(C.TIER1_DIR, "context", os.path.basename(r["context_file"])),
                     os.path.join(pkg, "tier1", "context", os.path.basename(r["context_file"])))
    for r in t2:
        shutil.copy2(os.path.join(C.TIER2_DIR, "packets", os.path.basename(r["packet_file"])),
                     os.path.join(pkg, "tier2", "packets", os.path.basename(r["packet_file"])))

    p1 = os.path.join(pkg, "tier1", "FILL_ME_tier1_reviewer_%s.csv" % tag)
    p2 = os.path.join(pkg, "tier2", "FILL_ME_tier2_reviewer_%s.csv" % tag)
    with open(p1, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(t1[0].keys()))
        w.writeheader()
        w.writerows(t1)
    with open(p2, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(t2[0].keys()))
        w.writeheader()
        w.writerows(t2)
    return pkg, t1, t2


def build_workbook(pkg, tag, t1, t2):
    """The same workbook shape eval/reviewer_xlsx_v3 produces, over the sampled rows."""
    wb = Workbook()
    wb.remove(wb.active)
    X._guide(wb, tag, len(t1), len(t2))

    ws1 = wb.create_sheet("Tier1")
    h1 = X.T1_KEEP + ["context_path"] + X.T1_FILL
    X._sheet(ws1, h1, len(X.T1_KEEP) + 1,
             [26, 26, 18, 16, 78] + [44, 30, 30, 34, 26, 30, 30, 26, 30, 12, 30])
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    missing = []
    for i, r in enumerate(t1, 2):
        rel = "tier1/" + r["context_file"]
        if not os.path.isfile(os.path.join(pkg, rel)):
            missing.append(rel)
        for j, k in enumerate(X.T1_KEEP, 1):
            ws1.cell(row=i, column=j, value=r.get(k, ""))
        # A path relative to the workbook, not an absolute one. The absolute form named this
        # machine's directories, which do not exist on the annotator's computer, so every link
        # in a shipped workbook was dead on arrival.
        c = ws1.cell(row=i, column=len(X.T1_KEEP) + 1, value=rel)
        c.hyperlink = rel
    dv1 = DataValidation(type="list", formula1='"' + ",".join(X.T1_CONFIDENCE) + '"',
                         allow_blank=True)
    ws1.add_data_validation(dv1)
    col = get_column_letter(len(h1) - 1)
    dv1.add(f"{col}2:{col}{len(t1) + 1}")

    ws2 = wb.create_sheet("Tier2")
    axes = list(C.TIER2_LABEL_AXES)
    h2 = X.T2_KEEP + ["packet_path"] + axes + ["notes"]
    X._sheet(ws2, h2, len(X.T2_KEEP) + 1, [22, 16, 44, 12, 78] + [20, 32, 20, 14, 30, 40])
    for i, r in enumerate(t2, 2):
        rel = "tier2/" + r["packet_file"]
        if not os.path.isfile(os.path.join(pkg, rel)):
            missing.append(rel)
        for j, k in enumerate(X.T2_KEEP, 1):
            ws2.cell(row=i, column=j, value=r.get(k, ""))
        c = ws2.cell(row=i, column=len(X.T2_KEEP) + 1, value=rel)
        c.hyperlink = rel
    for k, ax in enumerate(axes):
        dv = DataValidation(type="list", formula1='"' + ",".join(C.TIER2_LABEL_DOMAINS[ax]) + '"',
                            allow_blank=True, showErrorMessage=True,
                            errorTitle="Giá trị không hợp lệ",
                            error="Chỉ nhận các giá trị trong danh sách, hoặc để trống.")
        ws2.add_data_validation(dv)
        col = get_column_letter(len(X.T2_KEEP) + 2 + k)
        dv.add(f"{col}2:{col}{len(t2) + 1}")

    X._values_sheet(wb)
    X._metadata_sheet(wb, tag)
    if missing:
        raise SystemExit("%d row(s) point at a path that does not exist, refusing to save"
                         % len(missing))
    # A workbook is only shippable if no judgment cell carries a value. Check the saved object,
    # not the intent: this is the one guarantee the whole study rests on.
    fill_cols = set(X.T1_FILL) | set(C.TIER2_LABEL_AXES) | {"notes"}
    for ws, head in ((ws1, h1), (ws2, h2)):
        for j, name in enumerate(head, 1):
            if name in fill_cols:
                for i in range(2, ws.max_row + 1):
                    if (ws.cell(row=i, column=j).value or "") != "":
                        raise SystemExit(f"pre-filled judgment cell at {ws.title}!{name} row {i}")
    # The workbook lives inside the package, because its links are relative to it. Moving the
    # file out of this folder breaks every link, and the guide sheet says so.
    out = os.path.join(pkg, "reviewer_%s.xlsx" % tag)
    wb.save(out)
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--target", type=int, default=200)
    ap.add_argument("--sample-only", action="store_true")
    # A third blind annotator is the cheapest strengthening left: the sample, the packets and the
    # sealed key already exist, so adding one is a package build rather than a new study. Kept as a
    # flag so the same draw is reused and the new reader sees exactly what A and B saw.
    ap.add_argument("--reviewers", default="A,B",
                    type=lambda x: tuple(t.strip() for t in x.split(",") if t.strip()))
    ap.add_argument("--dest", default=os.path.join(C.ADJ_DIR, "SEND-HUMAN-2026-08-10"))
    a = ap.parse_args()
    if not TARGET_MIN <= a.target <= TARGET_MAX:
        sys.exit(f"--target {a.target} is outside the protocol's {TARGET_MIN} to {TARGET_MAX}")

    units, picked, strata, cuts, sizes = draw(a.target)
    key = C.read_json(os.path.join(C.TIER2_DIR, "BLINDING_KEY.json"))["payload"]["map"]
    uid2tool = {e["finding_uid"]: e["tool"] for e in key.values()}
    per_tool = Counter(uid2tool.get(u["finding_uid"], "?") for u in picked)
    records = sorted({(u["slug"], u["cve"]) for u in picked})

    payload = {
        "schema_version": "defect-study-sample-v3",
        "script": "eval/defect_study_v3.py",
        "protocol": "revision-cns-v2/bundle-src/DEFECT-LEVEL-STUDY-PROTOCOL.md",
        "seed": SEED, "target": a.target,
        "population": {"source": os.path.relpath(POP, SYS_ROOT), "topk": TOPK,
                       "n_findings": len(units),
                       "n_records": len({(u["slug"], u["cve"]) for u in units})},
        "stratification": {
            "axes": ["advisory_class", "patch_shape", "plugin_size"],
            "patch_shape_rule": ("refactor-heavy if >= %d changed PHP files, else rename if any "
                                 "PHP file was renamed, else deletion if any was deleted, else "
                                 "guard-insertion if any changed PHP file only gained lines, else "
                                 "code-replacement" % REFACTOR_FILES),
            "patch_shapes_absent_from_this_population": sorted(
                {"guard-insertion", "code-replacement", "deletion", "refactor-heavy", "rename"}
                - {u["_shape"] for u in units}),
            "absence_note": ("a shape the population does not contain cannot be given a floor. "
                             "This field exists so the gap is on the record rather than inferred "
                             "from a table that simply has no such row."),
            "plugin_size_rule": "PHP files in the vulnerable archive, tertiles of the sampled slugs",
            "plugin_size_cuts": list(cuts),
            "n_strata": len(strata), "floor_per_stratum": 1},
        "sample": {"n_findings": len(picked), "n_records": len(records),
                   "finding_uids": [u["finding_uid"] for u in picked],
                   "records": ["%s|%s" % r for r in records]},
        "strata": strata,
        "diagnostics": {
            "note": ("the protocol stratifies on class, patch shape and plugin size, not on tool. "
                     "The per-tool counts below are a consequence of that design, reported here "
                     "so an imbalance is seen before labelling rather than after."),
            "per_tool": dict(sorted(per_tool.items())),
            "per_class": dict(sorted(Counter(u["advisory_class"] for u in picked).items())),
            "per_shape": dict(sorted(Counter(u["_shape"] for u in picked).items())),
            "per_size": dict(sorted(Counter(u["_size"] for u in picked).items()))},
        "timestamp_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "python_version": platform.python_version(),
    }
    json.dump(payload, open(SAMPLE_OUT, "w", encoding="utf-8"), indent=1, sort_keys=True)
    print("sample: %d findings over %d records, %d strata, seed %d"
          % (len(picked), len(records), len(strata), SEED))
    print("  per class: " + ", ".join(f"{k} {v}" for k, v in payload["diagnostics"]["per_class"].items()))
    print("  per shape: " + ", ".join(f"{k} {v}" for k, v in payload["diagnostics"]["per_shape"].items()))
    print("  per size : " + ", ".join(f"{k} {v}" for k, v in payload["diagnostics"]["per_size"].items()))
    print("  per tool : " + ", ".join(f"{k} {v}" for k, v in payload["diagnostics"]["per_tool"].items()))
    print("  written  " + os.path.relpath(SAMPLE_OUT, SYS_ROOT))
    if a.sample_only:
        return 0

    for tag in a.reviewers:
        pkg, t1, t2 = build_package(picked, a.dest, tag)
        xlsx = build_workbook(pkg, tag, t1, t2)
        print("reviewer %s: %d tier-1 rows, %d tier-2 rows -> %s"
              % (tag, len(t1), len(t2), os.path.relpath(xlsx, SYS_ROOT)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
