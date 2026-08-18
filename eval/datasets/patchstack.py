#!/usr/bin/env python3
"""Patchstack gold-set DATASET ADAPTER.

The ONLY module that knows about the Patchstack data layout (the vulnerable+patched
plugin zips and the patchstack*.xlsx label sheets). Evaluation logic
(`eval/recall.py`, `eval/localize.py`) is dataset-agnostic and imports `load_rows()`
from here. To add another gold source, write a sibling adapter exposing the same
`load_rows()` contract: a list of dicts containing at least ``slug``, ``cve``,
``type``, ``cls``, ``vuln_zip``, and ``patched_zip``.

Data location: env PS_DIR, else ``<parent-of-project>/patchstack_bugbounty``.
Both the original development layout and the public WISP-1108 Zenodo layout are
supported.  For the public release, extract ``plugins.zip`` next to
``WISP-1108-CVE-Dataset.csv`` and point PS_DIR at that directory.
"""
from __future__ import annotations
import csv, glob, hashlib, os

_HERE = os.path.dirname(os.path.abspath(__file__))            # eval/datasets
_PROJECT = os.path.dirname(os.path.dirname(_HERE))            # project root
PS_DIR = os.environ.get("PS_DIR") or os.path.join(
    os.path.dirname(_PROJECT), "patchstack_bugbounty")        # sibling of the repo

TYPE_MAP = {
    "SQL Injection": "sqli",
    "Cross Site Request Forgery (CSRF)": "csrf",
    "Cross Site Scripting (XSS)": "xss",
    "PHP Object Injection": "deserial",
    "Deserialization of untrusted data": "deserial",
    "Remote Code Execution (RCE)": "rce",
    "Arbitrary Code Execution": "rce",
    "Broken Access Control": "auth",
    "Broken Authentication": "auth",
    "Privilege Escalation": "auth",
    "Bypass Vulnerability": "auth",
    "Insecure Direct Object References (IDOR)": "auth",
    "Local File Inclusion": "lfi",
    "Path Traversal": "lfi",
    "Arbitrary File Download": "lfi",
    "Server Side Request Forgery (SSRF)": "ssrf",
    "Arbitrary File Upload": "upload",
    # ambiguous / not cleanly in WISP taxonomy -> "other" (honest: hard to credit)
    "Sensitive Data Exposure": "other",
    "CSV Injection": "other",
    "Arbitrary File Deletion": "other",
    "Multiple Vulnerabilities": "other",
}


def classify_type(t):
    """Map a Patchstack vulnerability-type label to a WISP class, robust to label
    variants (short names, full CWE phrasings, hyphen/spacing). Keyword-based so
    new batches with different wording still classify correctly."""
    if t in TYPE_MAP:               # exact known label wins
        return TYPE_MAP[t]
    s = (t or "").lower()
    if not s:
        return "other"
    if "sql" in s:
        return "sqli"
    if "object injection" in s or "deserial" in s:
        return "deserial"
    if "cross-site request forgery" in s or "cross site request forgery" in s or "csrf" in s:
        return "csrf"
    if "server side request forgery" in s or "server-side request forgery" in s or "ssrf" in s:
        return "ssrf"
    if ("cross-site scripting" in s or "cross site scripting" in s or "xss" in s
            or "web page generation" in s):
        return "xss"
    if ("missing authorization" in s or "broken access" in s or "authentication" in s
            or "privilege" in s or "bypass" in s or "idor" in s
            or "insecure direct object" in s or "authorization" in s):
        return "auth"
    if "file upload" in s or "upload" in s:
        return "upload"
    if ("local file inclusion" in s or "path traversal" in s or "directory traversal" in s
            or "file download" in s or "file inclusion" in s or "lfi" in s):
        return "lfi"
    if ("remote code execution" in s or "arbitrary code" in s or "command injection" in s
            or "code injection" in s or "rce" in s):
        return "rce"
    return "other"


class DatasetSchemaError(ValueError):
    """The metadata cannot identify an exact vulnerable/patched archive pair."""


def _text(value):
    return "" if value is None else str(value).strip()


def _resolve(rel, data_dir=None):
    """Resolve a recorded 'plugins/<slug>/<file>.zip' path against the real tree.

    The xlsx records paths under 'plugins/', but batches are downloaded into
    'plugins_01/', 'plugins_02/', ... — search every 'plugins*' batch dir so the
    harness keeps working as new batches land.
    """
    data_dir = data_dir or PS_DIR
    rel = _text(rel).replace("\\", "/")
    if not rel:
        return ""
    cand = os.path.join(data_dir, rel)
    if os.path.exists(cand):
        return cand
    parts = rel.split("/", 1)
    if len(parts) == 2 and parts[0].startswith("plugins"):
        for base in sorted(glob.glob(os.path.join(data_dir, "plugins*"))):
            alt = os.path.join(base, parts[1])
            if os.path.exists(alt):
                return alt
        # The public archive collapses development batches (plugins_01, ...)
        # into one plugins/ directory while retaining the metadata path.
        public = os.path.join(data_dir, "plugins", parts[1])
        if os.path.exists(public):
            return public
    return cand


def sha256_file(path):
    """Return the SHA-256 of a regular file, or an empty string if it is absent."""
    if not path or not os.path.isfile(path):
        return ""
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def archive_provenance(row):
    """Stable, serializable archive identity for an evaluation result row."""
    vulnerable_hash = sha256_file(row.get("vuln_zip"))
    patched_hash = sha256_file(row.get("patched_zip"))
    expected_vulnerable = _text(row.get("vuln_sha256"))
    expected_patched = _text(row.get("patched_sha256"))
    return {
        "vulnerable_file": _text(row.get("vuln_file")),
        "patched_file": _text(row.get("patched_file")),
        "vulnerable_sha256": vulnerable_hash,
        "patched_sha256": patched_hash,
        "vulnerable_hash_matches_metadata": (
            not expected_vulnerable or vulnerable_hash == expected_vulnerable),
        "patched_hash_matches_metadata": (
            not expected_patched or patched_hash == expected_patched),
    }


def _record(row, idx, data_dir, metadata_file):
    required = ("Slug", "CVE", "Vulnerability Type", "Vulnerable File", "Patched File")
    missing = [name for name in required if name not in idx]
    if missing:
        raise DatasetSchemaError(
            f"{metadata_file} lacks required archive-identity field(s): {', '.join(missing)}")
    vuln_type = row[idx["Vulnerability Type"]]
    vuln_file = _text(row[idx["Vulnerable File"]])
    patched_file = _text(row[idx["Patched File"]])
    return {
        "slug": _text(row[idx["Slug"]]),
        "cve": _text(row[idx["CVE"]]),
        "type": _text(vuln_type),
        "cls": classify_type(vuln_type),
        "vuln_file": vuln_file,
        "patched_file": patched_file,
        "vuln_zip": _resolve(vuln_file, data_dir),
        "patched_zip": _resolve(patched_file, data_dir),
        "vuln_version": _text(row[idx["Vulnerable Version"]])
            if "Vulnerable Version" in idx else "",
        "patched_version": _text(row[idx["Patched Version"]])
            if "Patched Version" in idx else "",
        "metadata_file": os.path.abspath(metadata_file),
    }


def _legacy_rows(data_dir):
    """Read the authors' original Patchstack workbook layout."""
    import openpyxl
    out = []
    files = sorted(glob.glob(os.path.join(data_dir, "patchstack*.xlsx")),
                   key=lambda f: -os.path.getsize(f))
    for xf in files:
        wb = openpyxl.load_workbook(xf, data_only=True, read_only=True)
        if "Vulnerable Plugins" not in wb.sheetnames:
            continue
        rows = list(wb["Vulnerable Plugins"].iter_rows(values_only=True))
        if not rows:
            continue
        idx = {h: i for i, h in enumerate(rows[0])}
        for row in rows[1:]:
            out.append(_record(row, idx, data_dir, xf))
        wb.close()
    return out


def _public_csv(data_dir):
    candidates = [
        os.path.join(data_dir, "WISP-1108-CVE-Dataset.csv"),
        os.path.join(data_dir, "metadata", "WISP-1108-CVE-Dataset.csv"),
    ]
    return next((path for path in candidates if os.path.isfile(path)), "")


def _archive_manifest(data_dir):
    """Load the immutable archive-pair manifest used by corrected public releases."""
    candidates = [
        os.path.join(data_dir, "archive-manifest.csv"),
        os.path.join(data_dir, "archive_manifest.csv"),
        os.path.join(data_dir, "metadata", "archive-manifest.csv"),
        os.path.join(data_dir, "metadata", "archive_manifest.csv"),
    ]
    path = next((candidate for candidate in candidates if os.path.isfile(candidate)), "")
    if not path:
        return {}, ""
    manifest = {}
    with open(path, newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            key = (_text(row.get("Slug")), _text(row.get("CVE")))
            if not all(key):
                raise DatasetSchemaError(f"{path} contains a row with a blank Slug/CVE key")
            if key in manifest:
                raise DatasetSchemaError(f"{path} contains duplicate identity {key!r}")
            manifest[key] = row
    return manifest, path


def _public_rows(data_dir):
    """Read the machine-readable CSV shipped by Zenodo DOI 21627535."""
    path = _public_csv(data_dir)
    if not path:
        return []
    out = []
    manifest, manifest_path = _archive_manifest(data_dir)
    with open(path, newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            key = (_text(row.get("Slug")), _text(row.get("CVE")))
            identity = manifest.get(key, {})
            vuln_file = _text(identity.get("Vulnerable File") or row.get("Vulnerable File"))
            patched_file = _text(identity.get("Patched File") or row.get("Patched File"))
            if not patched_file:
                raise DatasetSchemaError(
                    f"{path} has no Patched File for {key!r}; use a corrected dataset "
                    "release or provide metadata/archive-manifest.csv")
            vuln_type = row.get("Vulnerability Type") or row.get("Class (full)") or ""
            out.append({
                "slug": key[0], "cve": key[1], "type": _text(vuln_type),
                "cls": _text(row.get("Class")) or classify_type(vuln_type),
                "vuln_file": vuln_file, "patched_file": patched_file,
                "vuln_zip": _resolve(vuln_file, data_dir),
                "patched_zip": _resolve(patched_file, data_dir),
                "vuln_version": _text(row.get("Vulnerable Version")),
                "patched_version": _text(row.get("Patched Version")),
                "vuln_sha256": _text(identity.get("Vulnerable SHA256")),
                "patched_sha256": _text(identity.get("Patched SHA256")),
                "metadata_file": os.path.abspath(manifest_path or path),
            })
    if manifest:
        loaded = {(row["slug"], row["cve"]) for row in out}
        extra = sorted(set(manifest) - loaded)
        if extra:
            raise DatasetSchemaError(
                f"{manifest_path} contains {len(extra)} identities absent from {path}")
    return out


def load_rows(data_dir=None):
    """Union every patchstack*.xlsx in PS_DIR, deduped by (slug, cve). Scales as
    new labeled batches land (patchstack_plugins_02.xlsx, ...). Every returned
    record carries the exact vulnerable and patched ZIP named by that metadata
    row; directory scans and filename guessing are deliberately forbidden."""
    # Standalone reproduction path (reviewer 7): the re-analyses only read slug, cve
    # and cls per record, so a shipped corpus index replaces the full xlsx + archive
    # corpus. When WISP_REPRO_INDEX points at that JSON we return it verbatim.
    idx = os.environ.get("WISP_REPRO_INDEX")
    if idx:
        import json
        with open(idx, encoding="utf-8") as handle:
            return json.load(handle)
    data_dir = os.path.abspath(data_dir or PS_DIR)
    out, seen = [], {}
    rows = _legacy_rows(data_dir)
    if not rows:
        rows = _public_rows(data_dir)
    for row in rows:
        key = (row["slug"], row["cve"])
        if not all(key):
            raise DatasetSchemaError(f"blank dataset identity in {row.get('metadata_file', data_dir)}")
        if key in seen:
            previous = seen[key]
            previous_pair = tuple(os.path.basename(previous[name])
                                  for name in ("vuln_zip", "patched_zip"))
            current_pair = tuple(os.path.basename(row[name])
                                 for name in ("vuln_zip", "patched_zip"))
            if previous_pair != current_pair:
                raise DatasetSchemaError(
                    f"conflicting archive pair for {key!r}: {previous_pair!r} != {current_pair!r}")
            continue
        seen[key] = row
        out.append(row)
    return out
